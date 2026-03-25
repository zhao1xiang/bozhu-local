"""
生成患者批量导入Excel模板
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_patient_import_template():
    """创建患者导入模板（含预约字段）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者信息"
    
    headers = [
        # 患者字段
        ("姓名", "必填"),
        ("门诊号", "选填"),
        ("就诊卡号", "选填"),
        ("手机号", "选填，11位数字"),
        ("诊断", "选填，多个用逗号分隔"),
        ("诊断其他说明", "选填"),
        ("药物类型", "选填，多个用逗号分隔"),
        ("药物其他说明", "选填"),
        ("左眼裸眼视力", "选填"),
        ("右眼裸眼视力", "选填"),
        ("左眼矫正视力", "选填"),
        ("右眼矫正视力", "选填"),
        ("左眼注射", "选填，是/否"),
        ("右眼注射", "选填，是/否"),
        ("患者类型", "选填，初治/经治"),
        ("已完成针数", "选填，仅经治患者填写"),
        ("备注", "选填"),
        # 预约字段（可选，不填则不创建预约）
        ("预约日期", "选填，格式 2024-01-01"),
        ("预约眼别", "选填，左眼/右眼/双眼"),
        ("预约药品", "选填"),
        ("注药医生", "选填"),
        ("费别", "选填"),
        ("针次", "选填，数字"),
        ("治疗阶段", "选填，强化期/巩固期"),
        ("时间段", "选填，上午/下午"),
        ("状况", "选填，稳定/不稳定"),
        ("预约备注", "选填"),
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, (header, desc) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # 第二行写入说明
        desc_cell = ws.cell(row=2, column=col_idx)
        desc_cell.value = desc
        desc_cell.font = Font(size=9, color="666666", italic=True)
        desc_cell.alignment = Alignment(horizontal='center', vertical='center')
        desc_cell.border = border
    
    # 添加示例数据
    example_data = [
        "张三", "MZ001", "JZK001", "13800138000", "湿性AMD,糖尿病视网膜病变", "", "雷珠单抗,康柏西普", "",
        "0.5", "0.6", "0.8", "0.9", "是", "否", "初治", "", "这是备注信息",
        "2024-06-01", "右眼", "雷珠单抗", "张医生", "自费", "1", "强化期", "上午", "稳定", "",
    ]
    
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    column_widths = [12, 12, 12, 15, 15, 15, 15, 15, 12, 12, 12, 12, 12, 12, 12, 12, 20,
                     14, 12, 12, 12, 10, 8, 12, 10, 12, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # 冻结前两行
    ws.freeze_panes = 'A3'
    
    # 保存文件
    wb.save("患者批量导入模板.xlsx")
    print("模板已生成：患者批量导入模板.xlsx")

if __name__ == "__main__":
    create_patient_import_template()
